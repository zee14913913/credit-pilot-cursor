# reminder_system.py
# CreditPilot - 定时提醒系统（每晚10点）

from sqlalchemy.orm import Session
from datetime import date, datetime, timedelta
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from pathlib import Path

from models import Statement, ReminderLog
from email_service import send_reminder_email

REPORTS_DIR = Path(os.getenv("REPORTS_DIR", "/app/reports"))
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# 收件人邮箱（小助理）
RECIPIENT_EMAIL = os.getenv("RECIPIENT_EMAIL", "wang041396@gmail.com")


def generate_daily_reminder(db: Session) -> Dict:
    """
    生成每日提醒消息
    
    Returns:
        {
            "tomorrow": List[Dict],
            "day_after": List[Dict],
            "total_gz_payment": float,
            "total_owner_payment": float,
            "most_urgent_client": str
        }
    """
    today = date.today()
    tomorrow = today + timedelta(days=1)
    day_after = today + timedelta(days=2)
    
    # 查询明天到期的账单
    tomorrow_statements = db.query(Statement).filter(
        Statement.is_active == True,
        Statement.due_date == tomorrow
    ).all()
    
    # 查询后天到期的账单
    day_after_statements = db.query(Statement).filter(
        Statement.is_active == True,
        Statement.due_date == day_after
    ).all()
    
    def format_statement_info(s: Statement):
        """格式化账单信息"""
        gz_payment = s.gz_payment_1 or 0.0
        owner_payment = s.owner_payment or 0.0
        doc_count = len(s.documents)
        
        # 判断付款状态
        gz_status = "✓ 已付" if gz_payment > 0 else "⚠️ 待付款"
        owner_status = "✓ 已付" if owner_payment > 0 else "⚠️ 待付款"
        doc_status = f"{doc_count}/3 ✓" if doc_count >= 3 else f"{doc_count}/3 ⚠️"
        
        return {
            "client_name": s.client_name,
            "card_number": s.card_number,
            "bank_name": s.bank_name,
            "gz_payment": gz_payment,
            "owner_payment": owner_payment,
            "gz_status": gz_status,
            "owner_status": owner_status,
            "document_count": doc_count,
            "doc_status": doc_status,
            "is_verified": s.is_verified
        }
    
    tomorrow_list = [format_statement_info(s) for s in tomorrow_statements]
    day_after_list = [format_statement_info(s) for s in day_after_statements]
    
    # 计算总额
    all_statements = tomorrow_statements + day_after_statements
    total_gz_payment = sum(s.gz_payment_1 or 0.0 for s in all_statements)
    total_owner_payment = sum(s.owner_payment or 0.0 for s in all_statements)
    
    # 计算最紧急客户（优先级：GZ代付 > Owner付款 > 单据缺失）
    most_urgent = calculate_most_urgent_client(all_statements)
    
    result = {
        "tomorrow": tomorrow_list,
        "day_after": day_after_list,
        "tomorrow_count": len(tomorrow_list),
        "day_after_count": len(day_after_list),
        "total_gz_payment": round(total_gz_payment, 2),
        "total_owner_payment": round(total_owner_payment, 2),
        "most_urgent_client": most_urgent
    }
    
    # 保存到日志
    log = ReminderLog(
        reminder_date=today,
        reminder_time="22:00",
        tomorrow_count=len(tomorrow_list),
        day_after_count=len(day_after_list),
        total_gz_payment=total_gz_payment,
        total_owner_payment=total_owner_payment,
        most_urgent_client=most_urgent.get("client_name", "") if most_urgent else ""
    )
    db.add(log)
    db.commit()
    
    return result


def calculate_most_urgent_client(statements: List[Statement]) -> Dict:
    """计算最紧急客户"""
    if not statements:
        return {}
    
    max_priority = -1
    most_urgent = None
    
    for s in statements:
        priority = 0
        
        # GZ需代付（优先级最高）
        if s.gz_payment_1 == 0 or s.gz_payment_1 is None:
            priority += 100
        
        # Owner需付款
        if s.owner_payment == 0 or s.owner_payment is None:
            priority += 50
        
        # 单据缺失
        doc_count = len(s.documents)
        if doc_count < 3:
            priority += (3 - doc_count) * 10
        
        # 未验证
        if not s.is_verified:
            priority += 5
        
        if priority > max_priority:
            max_priority = priority
            most_urgent = s
    
    if not most_urgent:
        return {}
    
    reasons = []
    if (most_urgent.gz_payment_1 == 0 or most_urgent.gz_payment_1 is None):
        reasons.append(f"GZ需代付 RM {most_urgent.gz_expenses:.2f}")
    if (most_urgent.owner_payment == 0 or most_urgent.owner_payment is None):
        reasons.append(f"Owner需付款 RM {most_urgent.owner_expenses:.2f}")
    
    doc_count = len(most_urgent.documents)
    if doc_count < 3:
        reasons.append(f"缺少{3 - doc_count}份单据")
    
    total_amount = (most_urgent.gz_payment_1 or 0.0) + (most_urgent.owner_payment or 0.0)
    
    return {
        "client_name": most_urgent.client_name,
        "card_number": most_urgent.card_number,
        "total_amount": round(total_amount, 2),
        "reasons": reasons
    }


def generate_excel_report(db: Session) -> str:
    """生成Excel日报"""
    today = date.today()
    reminder_data = generate_daily_reminder(db)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "每日提醒报告"
    
    # 标题样式
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 写入标题
    ws['A1'] = "CreditPilot 每日提醒报告"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"生成日期: {today.strftime('%Y-%m-%d')}"
    ws['A3'] = f"生成时间: {datetime.now().strftime('%H:%M:%S')}"
    
    # 明天到期
    row = 5
    ws[f'A{row}'] = f"明天到期 ({reminder_data['tomorrow_count']} 笔)"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    headers = ["客户", "卡号", "GZ代付", "Owner付款", "单据状态"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for stmt in reminder_data['tomorrow']:
        row += 1
        ws.cell(row=row, column=1, value=stmt['client_name'])
        ws.cell(row=row, column=2, value=stmt['card_number'])
        ws.cell(row=row, column=3, value=f"RM {stmt['gz_payment']:.2f}")
        ws.cell(row=row, column=4, value=f"RM {stmt['owner_payment']:.2f}")
        ws.cell(row=row, column=5, value=stmt['doc_status'])
    
    # 后天到期
    row += 2
    ws[f'A{row}'] = f"后天到期 ({reminder_data['day_after_count']} 笔)"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for stmt in reminder_data['day_after']:
        row += 1
        ws.cell(row=row, column=1, value=stmt['client_name'])
        ws.cell(row=row, column=2, value=stmt['card_number'])
        ws.cell(row=row, column=3, value=f"RM {stmt['gz_payment']:.2f}")
        ws.cell(row=row, column=4, value=f"RM {stmt['owner_payment']:.2f}")
        ws.cell(row=row, column=5, value=stmt['doc_status'])
    
    # 汇总
    row += 2
    ws[f'A{row}'] = "总计"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    ws[f'A{row}'] = f"GZ代付总额: RM {reminder_data['total_gz_payment']:.2f}"
    row += 1
    ws[f'A{row}'] = f"Owner付款总额: RM {reminder_data['total_owner_payment']:.2f}"
    row += 1
    ws[f'A{row}'] = f"合计: RM {reminder_data['total_gz_payment'] + reminder_data['total_owner_payment']:.2f}"
    
    # 最紧急客户
    if reminder_data['most_urgent_client']:
        row += 2
        urgent = reminder_data['most_urgent_client']
        ws[f'A{row}'] = f"🔴 最紧急: {urgent['client_name']}"
        ws[f'A{row}'].font = Font(bold=True, color="FF0000")
        row += 1
        ws[f'A{row}'] = f"金额: RM {urgent['total_amount']:.2f}"
        row += 1
        ws[f'A{row}'] = f"原因: {', '.join(urgent['reasons'])}"
    
    # 调整列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # 保存文件
    filename = f"CreditPilot_Daily_Report_{today.strftime('%Y%m%d')}.xlsx"
    filepath = REPORTS_DIR / filename
    wb.save(filepath)
    
    # 更新日志
    log = db.query(ReminderLog).filter(
        ReminderLog.reminder_date == today
    ).order_by(ReminderLog.id.desc()).first()
    
    if log:
        log.excel_report_path = str(filepath)
        db.commit()
    
    # 发送邮件
    try:
        send_reminder_email(
            recipient_email=RECIPIENT_EMAIL,
            excel_file_path=str(filepath),
            reminder_data=reminder_data
        )
    except Exception as e:
        print(f"⚠️ 邮件发送失败（但Excel已生成）: {e}")
    
    return str(filepath)


def print_reminder_message(reminder_data: Dict):
    """打印提醒消息到控制台"""
    print("=" * 60)
    print("📋 CreditPilot 明日到期提醒")
    print("=" * 60)
    print()
    
    if reminder_data['tomorrow_count'] > 0:
        print(f"明天到期 ({reminder_data['tomorrow'][0]['client_name']} 等 {reminder_data['tomorrow_count']} 笔):")
        for i, stmt in enumerate(reminder_data['tomorrow'], 1):
            print(f"{i}️⃣ {stmt['client_name']} - {stmt['bank_name']} *{stmt['card_number']}")
            print(f"   💰 GZ Pay: RM {stmt['gz_payment']:.2f} {stmt['gz_status']}")
            print(f"   💰 Owner Pay: RM {stmt['owner_payment']:.2f} {stmt['owner_status']}")
            print(f"   📎 单据: {stmt['doc_status']}")
            print()
    
    if reminder_data['day_after_count'] > 0:
        print(f"后天到期 ({reminder_data['day_after_count']} 笔):")
        for i, stmt in enumerate(reminder_data['day_after'], 1):
            print(f"{i}️⃣ {stmt['client_name']} - {stmt['bank_name']} *{stmt['card_number']}")
            print(f"   💰 GZ Pay: RM {stmt['gz_payment']:.2f} {stmt['gz_status']}")
            print(f"   💰 Owner Pay: RM {stmt['owner_payment']:.2f} {stmt['owner_status']}")
            print()
    
    print("-" * 60)
    print(f"总计需代付 (GZ): RM {reminder_data['total_gz_payment']:.2f}")
    print(f"总计需客户付 (Owner): RM {reminder_data['total_owner_payment']:.2f}")
    print(f"合计: RM {reminder_data['total_gz_payment'] + reminder_data['total_owner_payment']:.2f}")
    print()
    
    if reminder_data['most_urgent_client']:
        urgent = reminder_data['most_urgent_client']
        print(f"🔴 最紧急: {urgent['client_name']} (RM {urgent['total_amount']:.2f})")
        print(f"   原因: {' | '.join(urgent['reasons'])}")
    
    print("=" * 60)


def setup_scheduler():
    """设置定时调度器（每晚10点执行）"""
    from apscheduler.schedulers.blocking import BlockingScheduler
    from apscheduler.triggers.cron import CronTrigger
    
    scheduler = BlockingScheduler()
    
    def daily_reminder_job():
        """定时任务：生成每日提醒"""
        from models import SessionLocal
        db = SessionLocal()
        try:
            reminder_data = generate_daily_reminder(db)
            print_reminder_message(reminder_data)
            report_path = generate_excel_report(db)
            print(f"✓ Excel日报已生成: {REPORTS_DIR}")
            print(f"✓ 邮件已发送到: {RECIPIENT_EMAIL}")
        except Exception as e:
            print(f"✗ 提醒任务失败: {e}")
        finally:
            db.close()
    
    # 每晚22:00执行
    scheduler.add_job(
        daily_reminder_job,
        trigger=CronTrigger(hour=22, minute=0),
        id='daily_reminder',
        name='每日提醒任务',
        replace_existing=True
    )
    
    print("=" * 60)
    print("CreditPilot 定时提醒系统已启动")
    print("每晚22:00执行提醒任务")
    print("=" * 60)
    
    scheduler.start()


if __name__ == "__main__":
    # 立即测试（不等到晚上10点）
    from models import SessionLocal
    db = SessionLocal()
    try:
        reminder_data = generate_daily_reminder(db)
        print_reminder_message(reminder_data)
        report_path = generate_excel_report(db)
        print(f"\n✓ Excel日报已生成: {report_path}")
        print(f"✓ 邮件已发送到: {RECIPIENT_EMAIL}")
    finally:
        db.close()
